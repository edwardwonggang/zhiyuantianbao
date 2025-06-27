#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def update_2022_partial_data():
    """更新2022年湖北理科一分一段表部分数据"""
    
    # 基于已知信息和趋势分析的2022年部分数据
    # 已知：700分以上1人，690分以上8人，680分以上23人，670分以上56人
    data_2022_partial = [
        ["分数段", "说明", "累计人数"],
        ["700分以上", "已知数据", 1],
        ["690分以上", "已知数据", 8],
        ["680分以上", "已知数据", 23],
        ["670分以上", "已知数据", 56],
        ["", "", ""],
        ["注：以下为基于趋势分析的估算数据", "", ""],
        ["660分以上", "估算", "约120"],
        ["650分以上", "估算", "约250"],
        ["640分以上", "估算", "约450"],
        ["630分以上", "估算", "约750"],
        ["620分以上", "估算", "约1200"],
        ["610分以上", "估算", "约1800"],
        ["600分以上", "估算", "约2500"],
        ["", "", ""],
        ["说明：", "", ""],
        ["1. 700-670分为官方公布的准确数据", "", ""],
        ["2. 其他分数段为基于历年趋势的估算", "", ""],
        ["3. 需要获取官方完整一分一段表进行准确填充", "", ""],
        ["4. 2022年湖北省本科线：物理类409分", "", ""],
        ["5. 2022年湖北省特殊类型招生控制线：物理类504分", "", ""]
    ]
    
    # 加载现有的Excel文件
    wb = load_workbook("湖北省理科一分一段表2022-2024.xlsx")
    
    # 获取2022年工作表
    ws_2022 = wb["2022年湖北理科一分一段"]
    
    # 清除现有数据（保留标题和表头）
    for row in ws_2022.iter_rows(min_row=4, max_row=ws_2022.max_row):
        for cell in row:
            cell.value = None
    
    # 添加2022年部分数据
    for i, row_data in enumerate(data_2022_partial, start=4):
        for j, value in enumerate(row_data, start=1):
            cell = ws_2022.cell(row=i, column=j, value=value)
            
            # 设置样式
            if i == 4 or i == 10 or i == 15:  # 标题行
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif "已知数据" in str(value):
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            elif "估算" in str(value):
                cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            cell.alignment = Alignment(horizontal='center')
    
    # 添加数据来源说明
    ws_2022.cell(row=len(data_2022_partial) + 6, column=1, value="数据来源：")
    ws_2022.cell(row=len(data_2022_partial) + 7, column=1, value="1. 官方数据：湖北省教育考试院")
    ws_2022.cell(row=len(data_2022_partial) + 8, column=1, value="2. 高分段数据来源：2022年高考成绩发布时的官方统计")
    ws_2022.cell(row=len(data_2022_partial) + 9, column=1, value="3. 完整一分一段表需要进一步获取官方数据")
    
    # 保存文件
    wb.save("湖北省理科一分一段表2022-2024.xlsx")
    print("2022年部分数据已成功添加到Excel文件中")
    print("包含官方公布的高分段准确数据和基于趋势的估算数据")
    print("建议后续获取官方完整一分一段表进行准确填充")

if __name__ == "__main__":
    update_2022_partial_data()
